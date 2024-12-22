#!/usr/bin/env python
# coding: utf-8

"""
gtfs_trips_hourly_reporter.py

This module processes General Transit Feed Specification (GTFS) data to generate
hourly reports of trips for selected routes and exports the results to an Excel workbook.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==============================
# CONFIGURATION SECTION - CUSTOMIZE HERE
# ==============================

# Input directory containing GTFS files
BASE_INPUT_PATH = r'\\your_file_path\here\\'

# Output directory for the Excel file
BASE_OUTPUT_PATH = r'\\your_file_path\here\\'

# GTFS files to load
gtfs_files = [
    'trips.txt',
    'stop_times.txt',
    'routes.txt',
    'stops.txt',
    'calendar.txt'
]

# Routes and directions to process
route_directions = [
    # Replace with your route name(s) and desired direction
    # Can be filtered to direction 0, 1, or None for no filter
    {'route_short_name': '310', 'direction_id': 0},   # Process only direction 0 for route 310
    {'route_short_name': '101', 'direction_id': None} # Process all directions for route 101
]

# ==============================
# END OF CONFIGURATION SECTION
# ==============================

def check_input_files(base_path, files):
    """Ensure all required GTFS files exist in the input directory."""
    if not os.path.exists(base_path):
        raise FileNotFoundError(f"The input directory {base_path} does not exist.")
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        if not os.path.exists(file_path):
            raise FileNotFoundError(
                f"The required GTFS file {file_name} does not exist in {base_path}."
            )

def load_gtfs_data(base_path, files):
    """Load GTFS files into Pandas DataFrames."""
    data = {}
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        data_name = file_name.replace('.txt', '')
        data[data_name] = pd.read_csv(file_path)
    return data

def fix_time_format(time_str):
    """
    Fix time formats by:
    - Adding leading zeros to single-digit hours
    - Converting hours greater than 23 by subtracting 24
    """
    parts = time_str.split(":")

    # Add leading zero if the hour is a single digit
    if len(parts[0]) == 1:
        parts[0] = '0' + parts[0]

    # Correct times where the hour exceeds 23 (indicating next day service)
    if int(parts[0]) >= 24:
        parts[0] = str(int(parts[0]) - 24).zfill(2)

    return ":".join(parts)

def process_and_export(data, route_directions, output_path):
    """Process the GTFS data and export trips per hour to an Excel workbook."""
    trips = data['trips']
    stop_times = data['stop_times']
    routes = data['routes']
    calendar = data['calendar']

    # Filter the calendar for services available on weekdays (Monday through Friday)
    relevant_service_ids = calendar[
        (calendar['monday'] == 1) & (calendar['tuesday'] == 1) &
        (calendar['wednesday'] == 1) & (calendar['thursday'] == 1) &
        (calendar['friday'] == 1)
    ]['service_id']

    print("Relevant service IDs for weekdays:")
    print(relevant_service_ids)

    # Filter trips to include only those with relevant service IDs
    trips_filtered = trips[trips['service_id'].isin(relevant_service_ids)]
    print("Filtered trips:")
    print(trips_filtered.head())

    # Merge stop_times with trips and routes
    merged_data = pd.merge(stop_times, trips_filtered, on='trip_id')
    merged_data = pd.merge(
        merged_data,
        routes[['route_id', 'route_short_name']],
        on='route_id'
    )

    # Apply time format correction
    merged_data['arrival_time'] = merged_data['arrival_time'].apply(fix_time_format)
    merged_data['departure_time'] = merged_data['departure_time'].apply(fix_time_format)

    # Convert to datetime.time objects
    merged_data['arrival_time'] = pd.to_datetime(
        merged_data['arrival_time'].str.strip(), format='%H:%M:%S'
    ).dt.time
    merged_data['departure_time'] = pd.to_datetime(
        merged_data['departure_time'].str.strip(), format='%H:%M:%S'
    ).dt.time

    print("Merged data:")
    print(merged_data.head())

    # Create a new Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Process each route and direction
    for rd in route_directions:
        route_short = rd['route_short_name']
        direction_id = rd['direction_id']

        if direction_id is not None:
            filtered = merged_data[
                (merged_data['route_short_name'] == route_short) &
                (merged_data['direction_id'] == direction_id)
            ]
        else:
            filtered = merged_data[merged_data['route_short_name'] == route_short]

        print(f"Filtered data for route {route_short} direction {direction_id}:")
        print(filtered.head())

        # Further filter to starting stops (stop_sequence == 1)
        start_times = filtered[filtered['stop_sequence'] == 1]
        print(f"Start times for route {route_short} direction {direction_id}:")
        print(start_times.head())

        # Extract hour from departure_time and count trips per hour
        start_times_hour = start_times.copy()
        start_times_hour['departure_hour'] = start_times_hour['departure_time'].apply(lambda t: t.hour)
        trips_per_hour = start_times_hour.groupby('departure_hour').size().reset_index(name='trip_count')
        print(f"Trips per hour for route {route_short} direction {direction_id}:")
        print(trips_per_hour)

        # Add a new sheet for the current route and direction
        sheet_name = (
            f"Route_{route_short}_Dir_{direction_id}" 
            if direction_id is not None 
            else f"Route_{route_short}_All_Dirs"
        )
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        ws.append(trips_per_hour.columns.tolist())

        # Write data rows
        for row in trips_per_hour.itertuples(index=False, name=None):
            ws.append(row)

        # Adjust column widths and alignments
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col) + 2
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length
            for cell in col:
                cell.alignment = Alignment(horizontal='center')

        print(f"Processed {sheet_name}")

    # Ensure output directory exists
    os.makedirs(output_path, exist_ok=True)

    # Save the workbook
    output_file = os.path.join(output_path, "Trips_Per_Hour_Selected_Routes.xlsx")
    wb.save(output_file)
    print("Trips per hour for selected routes successfully exported!")

def main():
    """Main function to execute the script."""
    try:
        # Check if all input files exist
        check_input_files(BASE_INPUT_PATH, gtfs_files)

        # Load GTFS data
        data = load_gtfs_data(BASE_INPUT_PATH, gtfs_files)

        # Process data and export to Excel
        process_and_export(data, route_directions, BASE_OUTPUT_PATH)

    except FileNotFoundError as fnf_error:
        print(f"File not found error: {fnf_error}")
    except pd.errors.ParserError as parse_error:
        print(f"Parsing error while reading GTFS files: {parse_error}")
    except PermissionError as perm_error:
        print(f"Permission error: {perm_error}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()

