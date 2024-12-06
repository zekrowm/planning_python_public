#!/usr/bin/env python
# coding: utf-8

# In[4]:


#!/usr/bin/env python
# coding: utf-8

import os
import pandas as pd
from datetime import timedelta
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==============================
# CONFIGURATION SECTION - CUSTOMIZE HERE
# ==============================

# Input directory containing GTFS files
gtfs_input_path = r'\\your_file_path\here\\'

# Output directory for the Excel file
output_path = r'\\your_file_path\here\\'

# GTFS files to load
gtfs_files = ['routes.txt', 'trips.txt', 'stop_times.txt', 'calendar.txt', 'calendar_dates.txt']

# Output Excel file name (base name)
output_excel = "route_schedule_headway_with_modes.xlsx"

# Define time blocks with start and end times in 'HH:MM' format
time_blocks_config = {
    'am': ('04:00', '09:00'),
    'midday': ('09:00', '15:00'),
    'pm': ('15:00', '21:00'),
    'night': ('21:00', '28:00')  # 28:00 is equivalent to 04:00 the next day
}

# Define multiple schedule types and their corresponding days
schedule_types = {
    'Weekday': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday'],
    'Saturday': ['saturday'],
    'Sunday': ['sunday'],
    # Add more if desired
}

# ==============================
# END OF CONFIGURATION SECTION
# ==============================

def check_input_files(base_path, files):
    if not os.path.exists(base_path):
        raise FileNotFoundError(f"The input directory {base_path} does not exist.")
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The required GTFS file {file_name} does not exist in {base_path}.")

def load_gtfs_data(base_path, files):
    data = {}
    for file_name in files:
        file_path = os.path.join(base_path, file_name)
        data_name = file_name.replace('.txt', '')
        try:
            data[data_name] = pd.read_csv(file_path)
            print(f"Loaded {file_name} with {len(data[data_name])} records.")
        except Exception as e:
            raise Exception(f"Error loading {file_name}: {e}")
    return data

def parse_time_blocks(time_blocks_str):
    parsed_blocks = {}
    for block_name, (start_str, end_str) in time_blocks_str.items():
        start_parts = start_str.split(':')
        end_parts = end_str.split(':')
        start_td = timedelta(hours=int(start_parts[0]), minutes=int(start_parts[1]))
        end_td = timedelta(hours=int(end_parts[0]), minutes=int(end_parts[1]))
        parsed_blocks[block_name] = (start_td, end_td)
    return parsed_blocks

def assign_time_block(time, blocks):
    for block_name, (start, end) in blocks.items():
        if start <= time < end:
            return block_name
    return 'other'

def format_timedelta(td):
    if pd.isna(td):
        return None
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{hours:02}:{minutes:02}"

def find_large_break(trip_times):
    late_morning = pd.Timedelta(hours=10)
    early_afternoon = pd.Timedelta(hours=14)
    midday_trips = trip_times[(trip_times >= late_morning) & (trip_times <= early_afternoon)]
    midday_trips = midday_trips.reset_index(drop=True)
    if len(midday_trips) < 2:
        return False
    for i in range(1, len(midday_trips)):
        if (midday_trips[i] - midday_trips[i - 1]) > pd.Timedelta(hours=3):
            return True
    return False

def calculate_trip_times(group):
    trip_times = group['departure_time'].sort_values()
    first_trip = trip_times.min()
    last_trip = trip_times.max()

    if first_trip >= pd.Timedelta(hours=15):
        # PM-only route
        return pd.Series({
            'first_trip_time': format_timedelta(first_trip),
            'last_trip_time': format_timedelta(last_trip),
            'am_last_trip_time': None,
            'pm_first_trip_time': format_timedelta(first_trip)
        })
    elif last_trip <= pd.Timedelta(hours=10):
        # AM-only route
        return pd.Series({
            'first_trip_time': format_timedelta(first_trip),
            'last_trip_time': format_timedelta(last_trip),
            'am_last_trip_time': format_timedelta(last_trip),
            'pm_first_trip_time': None
        })
    elif find_large_break(trip_times):
        # Normal route with midday break
        am_last_trip = trip_times[trip_times < pd.Timedelta(hours=10)].max()
        pm_first_trip = trip_times[trip_times > pd.Timedelta(hours=14)].min()
        return pd.Series({
            'first_trip_time': format_timedelta(first_trip),
            'last_trip_time': format_timedelta(last_trip),
            'am_last_trip_time': format_timedelta(am_last_trip),
            'pm_first_trip_time': format_timedelta(pm_first_trip)
        })
    else:
        # Normal all-day route
        return pd.Series({
            'first_trip_time': format_timedelta(first_trip),
            'last_trip_time': format_timedelta(last_trip),
            'am_last_trip_time': None,
            'pm_first_trip_time': None
        })

def calculate_headways(departure_times):
    sorted_times = departure_times.sort_values()
    headways = sorted_times.diff().dropna().apply(lambda x: x.total_seconds() / 60)
    if headways.empty:
        return None
    return headways.mode()[0]

def process_headways(merged_data, time_blocks):
    headways = merged_data.groupby(['route_short_name', 'route_long_name', 'direction_id', 'time_block'])['departure_time'].apply(calculate_headways).reset_index()
    headway_dict = {
        'weekday_am_headway': {},
        'weekday_midday_headway': {},
        'weekday_pm_headway': {},
        'weekday_night_headway': {}
    }
    for _, row in headways.iterrows():
        route = (row['route_short_name'], row['route_long_name'], row['direction_id'])
        if row['time_block'] == 'am':
            headway_dict['weekday_am_headway'][route] = row['departure_time']
        elif row['time_block'] == 'midday':
            headway_dict['weekday_midday_headway'][route] = row['departure_time']
        elif row['time_block'] == 'pm':
            headway_dict['weekday_pm_headway'][route] = row['departure_time']
        elif row['time_block'] == 'night':
            headway_dict['weekday_night_headway'][route] = row['departure_time']
    return headway_dict

def merge_headways(trip_times, headway_dict):
    trip_times['weekday_am_headway'] = trip_times.apply(
        lambda row: headway_dict['weekday_am_headway'].get((row['route_short_name'], row['route_long_name'], row['direction_id']), None),
        axis=1
    )
    trip_times['weekday_midday_headway'] = trip_times.apply(
        lambda row: headway_dict['weekday_midday_headway'].get((row['route_short_name'], row['route_long_name'], row['direction_id']), None),
        axis=1
    )
    trip_times['weekday_pm_headway'] = trip_times.apply(
        lambda row: headway_dict['weekday_pm_headway'].get((row['route_short_name'], row['route_long_name'], row['direction_id']), None),
        axis=1
    )
    trip_times['weekday_night_headway'] = trip_times.apply(
        lambda row: headway_dict['weekday_night_headway'].get((row['route_short_name'], row['route_long_name'], row['direction_id']), None),
        axis=1
    )
    return trip_times

def save_to_excel(final_data, output_path, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Route_Schedule_Headway"
    
    headers = final_data.columns.tolist()
    ws.append(headers)
    
    for row in final_data.itertuples(index=False, name=None):
        ws.append(row)
    
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length
        for cell in col:
            cell.alignment = Alignment(horizontal='center')
    
    os.makedirs(output_path, exist_ok=True)
    output_file_path = os.path.join(output_path, output_file)
    wb.save(output_file_path)
    print(f"Final data successfully saved to {output_file_path}")

def main():
    try:
        print("Checking input files...")
        check_input_files(gtfs_input_path, gtfs_files)
        print("All input files are present.\n")

        print("Loading GTFS data...")
        data = load_gtfs_data(gtfs_input_path, gtfs_files)
        print("GTFS data loaded successfully.\n")

        print("Parsing time block definitions...")
        time_blocks = parse_time_blocks(time_blocks_config)
        print("Time block definitions parsed.\n")

        # We now loop over each schedule type defined in schedule_types
        calendar = data['calendar']
        trips = data['trips']
        routes = data['routes']
        stop_times = data['stop_times']

        for schedule_type, days in schedule_types.items():
            print(f"Processing schedule: {schedule_type}")

            # Create a mask for services that run on all the specified days
            mask = pd.Series([True]*len(calendar))
            for day in days:
                mask &= (calendar[day] == 1)
            relevant_service_ids = calendar[mask]['service_id']

            if relevant_service_ids.empty:
                print(f"No services found for {schedule_type}. Skipping.\n")
                continue

            # Filter trips
            trips_filtered = trips[trips['service_id'].isin(relevant_service_ids)]
            if trips_filtered.empty:
                print(f"No trips found for {schedule_type}. Skipping.\n")
                continue

            # Merge routes and trip info
            trip_info = trips_filtered[['trip_id', 'route_id', 'service_id', 'direction_id']].merge(
                routes[['route_id', 'route_short_name', 'route_long_name']], on='route_id'
            )
            print(f"Merged trip information has {len(trip_info)} records for {schedule_type}.\n")

            # Merge trip info with stop_times
            merged_data = stop_times[['trip_id', 'departure_time', 'stop_sequence']].merge(
                trip_info, on='trip_id'
            )
            print(f"Merged data has {len(merged_data)} records for {schedule_type}.\n")

            # Filter to include only starting stops
            merged_data = merged_data[merged_data['stop_sequence'] == 1]
            print(f"Filtered starting trips count: {len(merged_data)} for {schedule_type}\n")

            if merged_data.empty:
                print(f"No starting trips for {schedule_type}. Skipping.\n")
                continue

            # Convert departure_time to timedelta
            merged_data['departure_time'] = pd.to_timedelta(merged_data['departure_time'], errors='coerce')
            merged_data = merged_data.dropna(subset=['departure_time'])
            if merged_data.empty:
                print(f"All departure_times invalid for {schedule_type}. Skipping.\n")
                continue

            print("Assigning time blocks...")
            merged_data['time_block'] = merged_data['departure_time'].apply(lambda x: assign_time_block(x, time_blocks))
            print("Time blocks assigned.\n")

            # Filter out 'other' time blocks
            merged_data = merged_data[merged_data['time_block'] != 'other']
            print(f"Trips after filtering 'other' time blocks: {len(merged_data)} for {schedule_type}\n")

            if merged_data.empty:
                print(f"No trips left after filtering 'other' time blocks for {schedule_type}. Skipping.\n")
                continue

            # Group by route and direction, calculate trip times
            print("Calculating trip times...")
            trip_times = merged_data.groupby(['route_short_name', 'route_long_name', 'direction_id']).apply(calculate_trip_times).reset_index()
            print("Trip times calculated.\n")

            # Calculate headways
            print("Calculating headways...")
            headway_dict = process_headways(merged_data, time_blocks)
            print("Headways calculated.\n")

            # Merge headways with trip times
            print("Merging headways with trip times...")
            final_data = merge_headways(trip_times, headway_dict)
            print("Headways merged.\n")

            # Save to Excel with schedule_type in filename
            output_file_for_schedule = f"{schedule_type}_{output_excel}"
            print(f"Saving data for {schedule_type} to Excel...")
            save_to_excel(final_data, output_path, output_file_for_schedule)
            print(f"Data for {schedule_type} saved.\n")

        print("All schedule types processed successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()

