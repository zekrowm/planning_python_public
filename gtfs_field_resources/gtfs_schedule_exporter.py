#!/usr/bin/env python
# coding: utf-8

from datetime import datetime, timedelta
import os
import re
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==============================
# CONFIGURATION SECTION - CUSTOMIZE HERE
# ==============================

# Input file paths for GTFS files (update these paths accordingly)
base_input_path = r"C:\Path\To\Your\System\GTFS_Data"  # Replace with your file path
trips_file = os.path.join(base_input_path, "trips.txt")
stop_times_file = os.path.join(base_input_path, "stop_times.txt")
routes_file = os.path.join(base_input_path, "routes.txt")
stops_file = os.path.join(base_input_path, "stops.txt")
calendar_file = os.path.join(base_input_path, "calendar.txt")

# Output directory (update this path accordingly)
base_output_path = r"C:\Path\To\Your\Output_Folder"  # Replace with your file path
if not os.path.exists(base_output_path):
    os.makedirs(base_output_path)

# List of route short names to process
# Set to 'all' (string) to include all routes,
# Or provide a list like ['101', '102', '103']
route_short_names_input = ['101', '102']  # Modify as needed

# Time format option: '24' for 24-hour time, '12' for 12-hour time
time_format_option = '12'  # Change to '24' for 24-hour format

# Placeholder values
missing_time = "---"

# Maximum column width for Excel output (used to wrap long headers)
max_column_width = 30  # Adjust as needed

# ==============================
# END OF CONFIGURATION SECTION
# ==============================

# ==============================
# UTILITY FUNCTIONS
# ==============================

def time_to_minutes(time_str):
    """
    Converts a time string to total minutes since midnight.
    Supports 'HH:MM' and 'HH:MM AM/PM' formats.
    Allows hours >=24 for 24-hour format.
    Returns None if the format is invalid.
    """
    if time_str == '---':
        return None
    try:
        # Match 12-hour and 24-hour formats
        match = re.match(r'^(\d{1,2}):(\d{2})(?:\s*(AM|PM))?$', time_str, re.IGNORECASE)
        if not match:
            return None
        hour, minute, period = match.groups()
        hour = int(hour)
        minute = int(minute)
        if period:
            period = period.upper()
            if period == 'PM' and hour != 12:
                hour += 12
            elif period == 'AM' and hour == 12:
                hour = 0
        # Allow hours >=24 by not constraining the hour value
        total_minutes = hour * 60 + minute
        return total_minutes
    except:
        return None

def check_schedule_order(df, ordered_stop_names, route_short_name, schedule_type, direction_id):
    """
    Checks that times in the DataFrame increase across rows and down columns, ignoring '---'.
    Prints warnings with emojis if violations are found, and a checkmark if the schedule passes.
    """
    violations = False

    # Row-wise check: Ensure times increase from left to right within each trip
    for idx, row in df.iterrows():
        last_time = None
        for stop in ordered_stop_names:
            time_str = row.get(f"{stop} Schedule", '---')
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                print(f"⚠️ Time order violation in Route '{route_short_name}', Schedule '{schedule_type}', Direction '{direction_id}', Trip '{row['Trip Headsign']}': '{stop}' time {time_str} is earlier than previous stop.")
                violations = True
                break  # Warn once per trip

    # Column-wise check: Ensure times increase from top to bottom within each stop
    for stop in ordered_stop_names:
        last_time = None
        for idx, row in df.iterrows():
            time_str = row.get(f"{stop} Schedule", '---')
            current_time = time_to_minutes(time_str)
            if current_time is None:
                continue
            if last_time is not None and current_time < last_time:
                print(f"⚠️ Time order violation in Route '{route_short_name}', Schedule '{schedule_type}', Direction '{direction_id}', Stop '{stop}': time {time_str} is earlier than previous trip.")
                violations = True
                break  # Warn once per stop

    if not violations:
        print("✅ Schedule order check passed.")

def adjust_time(time_str, time_format='24'):
    """
    Adjusts time strings to the desired format.
    If time_format is '24', keeps hours as is without wrapping.
    If time_format is '12', converts to 12-hour format with AM/PM, unless hours >=24.
    Returns None if the format is invalid.
    """
    parts = time_str.strip().split(":")
    if len(parts) >= 2:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])

            if time_format == '12':
                if hours >= 24:
                    # Cannot convert hours >=24 to 12-hour format meaningfully
                    print(f"Warning: Cannot convert time '{time_str}' to 12-hour format. Keeping as is.")
                    return time_str  # Or choose to set to '---' or another placeholder
                period = 'AM' if hours < 12 else 'PM'
                adjusted_hour = hours % 12
                if adjusted_hour == 0:
                    adjusted_hour = 12
                formatted_time = f"{adjusted_hour}:{minutes:02} {period}"
                return formatted_time
            else:
                # Keep hours as is for 24-hour format without wrapping
                return f"{hours:02}:{minutes:02}"
        except ValueError:
            print(f"Warning: Invalid time format encountered: '{time_str}'")
            return None
    else:
        print(f"Warning: Invalid time format encountered: '{time_str}'")
        return None

def get_ordered_stops(direction_id, relevant_trips):
    relevant_trips_direction = relevant_trips[relevant_trips['direction_id'] == direction_id]

    if relevant_trips_direction.empty:
        print(f"Warning: No trips found for direction_id '{direction_id}'.")
        return [], []

    all_stops = timepoints[timepoints['trip_id'].isin(relevant_trips_direction['trip_id'])]
    all_stops = all_stops.sort_values(['trip_id', 'stop_sequence'])

    if all_stops.empty:
        print(f"Warning: No stop times found for direction_id '{direction_id}'.")
        return [], []

    # Keep all occurrences of each stop (no drop_duplicates on stop_id)
    unique_stops = all_stops[['stop_id', 'stop_sequence']].drop_duplicates().sort_values('stop_sequence')

    stop_names = stops.set_index('stop_id')['stop_name']
    ordered_stop_names = [
        f"{stop_names.get(stop_id, f'Unknown Stop ID {stop_id}')} ({seq})"
        for stop_id, seq in zip(unique_stops['stop_id'], unique_stops['stop_sequence'])
    ]

    return ordered_stop_names, unique_stops

def map_service_id_to_schedule(service_row):
    """
    Maps a service_id row to a schedule type based on days served.
    Includes 'Weekday except Friday'.
    """
    days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    served_days = [day for day in days if service_row.get(day, '0') == '1']

    # Define day sets for different schedule types
    weekday = {'monday', 'tuesday', 'wednesday', 'thursday', 'friday'}
    weekday_except_friday = {'monday', 'tuesday', 'wednesday', 'thursday'}
    saturday = {'saturday'}
    sunday = {'sunday'}
    weekend = {'saturday', 'sunday'}
    daily = set(days)

    if not served_days:
        return 'Holiday'  # Or another appropriate label

    served_set = set(served_days)

    if served_set == weekday:
        return 'Weekday'
    elif served_set == weekday_except_friday:
        return 'Weekday_except_Friday'
    elif served_set == saturday:
        return 'Saturday'
    elif served_set == sunday:
        return 'Sunday'
    elif served_set == weekend:
        return 'Weekend'
    elif served_set == {'friday', 'saturday'}:
        return 'Friday-Saturday'
    elif served_set == daily:
        return 'Daily'
    else:
        return 'Special'  # For other combinations

def process_trips_for_direction(relevant_trips_direction, ordered_stop_names, unique_stops, time_format, route_short_name, schedule_type, direction_id):
    """
    Processes trips for a specific direction_id and returns a DataFrame without 'Trip ID'.
    Each stop occurrence is preserved, ensuring that repeated visits to the same stop appear as distinct columns.
    Sorts trips based on the latest departure time in 24-hour format.
    Checks for sequential departure times and prints warnings if inconsistencies are found.
    Also performs schedule order checks across rows and columns.
    """
    if relevant_trips_direction.empty:
        print("Warning: No trips to process for this direction.")
        return pd.DataFrame()

    ordered_stop_ids = unique_stops['stop_id'].tolist()
    ordered_stop_sequences = unique_stops['stop_sequence'].tolist()
    stop_index_map = {seq: i for i, seq in enumerate(ordered_stop_sequences)}

    output_data = []

    for trip_id, group in timepoints[timepoints['trip_id'].isin(relevant_trips_direction['trip_id'])].groupby('trip_id'):
        trip_info = relevant_trips_direction[relevant_trips_direction['trip_id'] == trip_id].iloc[0]
        route_name = routes[routes['route_id'] == trip_info['route_id']]['route_short_name'].values[0]
        trip_headsign = trip_info.get('trip_headsign', '')

        # Initialize the row with route_name, direction_id, trip_headsign
        row = [route_name, trip_info['direction_id'], trip_headsign]

        # Initialize schedule times with missing_time
        schedule_times = [missing_time] * len(ordered_stop_ids)
        valid_departure_times_24 = []

        for idx, stop in group.iterrows():
            departure_str = stop['departure_time'].strip()
            time_str_display = adjust_time(departure_str, time_format)
            time_str_24 = adjust_time(departure_str, '24')

            if time_str_display is None or time_str_24 is None:
                print(f"Warning: Invalid time format '{stop['departure_time']}' in trip_id '{trip_id}' at stop_id '{stop['stop_id']}'")
                continue

            seq = stop['stop_sequence']
            index = stop_index_map[seq]
            schedule_times[index] = time_str_display
            valid_departure_times_24.append(time_str_24)

        # Determine the sorting time based on the maximum departure time
        if valid_departure_times_24:
            try:
                # Use pd.to_timedelta which can handle hours >=24
                departure_timedeltas = [pd.to_timedelta(t + ':00') for t in valid_departure_times_24]
                max_sort_time = max(departure_timedeltas)
            except Exception as e:
                max_sort_time = pd.to_timedelta('00:00')
                print(f"Warning: Failed to determine maximum departure time for trip_id '{trip_id}'. Defaulting to '00:00'. Error: {e}")
        else:
            # If no valid times, assign a large timedelta to push this trip to the bottom
            max_sort_time = pd.to_timedelta('9999:00:00')

        # Add schedule times and the sort time to the row
        row.extend(schedule_times)
        row.append(max_sort_time)

        # Add this trip's data to the output
        output_data.append(row)

        # Check for sequential times within the trip (existing functionality)
        times_in_seconds = []
        for time_str in valid_departure_times_24:
            total_minutes = time_to_minutes(time_str)
            if total_minutes is not None:
                seconds = total_minutes * 60
                times_in_seconds.append(seconds)
            else:
                print(f"Warning: Failed to parse time '{time_str}' in trip_id '{trip_id}'.")

        for i in range(1, len(times_in_seconds)):
            if times_in_seconds[i] < times_in_seconds[i - 1]:
                print(
                    f"⚠️ Non-sequential departure times in trip_id '{trip_id}' for Route '{route_short_name}', "
                    f"Schedule '{schedule_type}', Direction '{direction_id}'. "
                    f"Stop {i + 1} is earlier than Stop {i}."
                )
                break  # Warn once per trip

    columns = ['Route Name', 'Direction ID', 'Trip Headsign'] + [f"{sn} Schedule" for sn in ordered_stop_names] + ['sort_time']
    df = pd.DataFrame(output_data, columns=columns)
    df = df.sort_values(by='sort_time').drop(columns=['sort_time'])

    # Perform schedule order check
    check_schedule_order(df, ordered_stop_names, route_short_name, schedule_type, direction_id)

    return df

def export_to_excel_multiple_sheets(df_dict, output_file):
    """
    Exports multiple DataFrames to an Excel file with each DataFrame in a separate sheet.
    """
    if not df_dict:
        print(f"No data to export to {output_file}.")
        return

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            if df.empty:
                print(f"No data for sheet '{sheet_name}'. Skipping...")
                continue
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Apply alignment to headers and adjust column widths
            for col_num, col_name in enumerate(df.columns, 1):  # 1-based indexing for openpyxl
                # Get the column letter
                col_letter = get_column_letter(col_num)

                # Set alignment to left and enable text wrapping for the header
                header_cell = worksheet[f'{col_letter}1']
                header_cell.alignment = Alignment(horizontal='left', wrap_text=True)

                # Set alignment to left for all data cells in the column
                for row_num in range(2, worksheet.max_row + 1):  # Start from row 2 to skip header
                    cell = worksheet[f'{col_letter}{row_num}']
                    cell.alignment = Alignment(horizontal='left')

                # Calculate the maximum width for this column
                column_cells = worksheet[col_letter]
                try:
                    max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                except:
                    max_length = 10  # Default width if calculation fails

                # Adjust the column width, limiting it to the maximum column width
                adjusted_width = min(max_length + 2, max_column_width)
                worksheet.column_dimensions[col_letter].width = adjusted_width

    print(f"Data exported to {output_file}")

# ==============================
# MAIN SCRIPT LOGIC
# ==============================

# Load GTFS files with basic error handling
try:
    trips = pd.read_csv(trips_file, dtype=str)
    stop_times = pd.read_csv(stop_times_file, dtype=str)
    routes = pd.read_csv(routes_file, dtype=str)
    stops = pd.read_csv(stops_file, dtype=str)
    calendar = pd.read_csv(calendar_file, dtype=str)
    print("Successfully loaded all GTFS files.")
except FileNotFoundError as e:
    print(f"Error: {e}")
    print("Please check your input file paths in the configuration section.")
    exit(1)
except Exception as e:
    print(f"An unexpected error occurred while reading GTFS files: {e}")
    exit(1)

# Convert 'stop_sequence' to numeric to ensure correct sorting
stop_times['stop_sequence'] = pd.to_numeric(stop_times['stop_sequence'], errors='coerce')
if stop_times['stop_sequence'].isnull().any():
    print("Warning: Some 'stop_sequence' values could not be converted to numeric.")

# Handle 'route_short_names_input' being 'all', a string, or a list
if isinstance(route_short_names_input, str):
    if route_short_names_input.lower() == 'all':
        route_short_names = routes['route_short_name'].dropna().unique().tolist()
        print(f"Selected all routes: {route_short_names}")
    else:
        # Assume comma-separated string
        route_short_names = [name.strip() for name in route_short_names_input.split(',')]
        print(f"Selected routes: {route_short_names}")
elif isinstance(route_short_names_input, list):
    route_short_names = route_short_names_input
    print(f"Selected routes: {route_short_names}")
else:
    print("Error: 'route_short_names_input' must be either 'all', a comma-separated string, or a list of route short names.")
    exit(1)

# Check for 'timepoint' column and filter timepoints
if 'timepoint' in stop_times.columns:
    timepoints = stop_times[stop_times['timepoint'] == '1']
    print("Filtered stop_times based on 'timepoint' column.")
else:
    print("Warning: 'timepoint' column not found. Using all stops as timepoints.")
    timepoints = stop_times.copy()

# Mapping service_id to schedule types
service_id_schedule_map = {}
schedule_types_set = set()

for _, service_row in calendar.iterrows():
    service_id = service_row['service_id']
    schedule_type = map_service_id_to_schedule(service_row)
    service_id_schedule_map[service_id] = schedule_type
    schedule_types_set.add(schedule_type)

print(f"Identified schedule types: {schedule_types_set}")

# Process each route and schedule_type
for route_short_name in route_short_names:
    print(f"\nProcessing route '{route_short_name}'...")

    # Get route_ids for the current route_short_name
    route_ids = routes[routes['route_short_name'] == route_short_name]['route_id']
    if route_ids.empty:
        print(f"Error: Route '{route_short_name}' not found in routes.txt.")
        continue  # Skip to next route

    # Process each schedule_type
    for schedule_type in schedule_types_set:
        print(f"  Processing schedule type '{schedule_type}'...")

        # Get service_ids for this schedule_type
        relevant_service_ids = [sid for sid, stype in service_id_schedule_map.items() if stype == schedule_type]

        if not relevant_service_ids:
            print(f"    No services found for schedule type '{schedule_type}'.")
            continue

        # Get trips for this route and schedule_type
        relevant_trips = trips[
            (trips['route_id'].isin(route_ids)) &
            (trips['service_id'].isin(relevant_service_ids))
        ]

        if relevant_trips.empty:
            print(f"    No trips found for route '{route_short_name}' with schedule type '{schedule_type}'.")
            continue

        # Get unique direction_ids within these trips
        direction_ids = relevant_trips['direction_id'].unique()

        # Dictionary to hold DataFrames for each direction_id
        df_sheets = {}

        for direction_id in direction_ids:
            print(f"    Processing direction_id '{direction_id}'...")

            # Get trips for this direction_id
            trips_direction = relevant_trips[relevant_trips['direction_id'] == direction_id]

            # Get ordered stops for this direction_id
            ordered_stop_names, ordered_stop_ids = get_ordered_stops(direction_id, trips_direction)

            if not ordered_stop_names:
                print(f"      No stops found for direction_id '{direction_id}'. Skipping...")
                continue

            # Process trips for this direction_id
            df = process_trips_for_direction(
                trips_direction,
                ordered_stop_names,
                ordered_stop_ids,
                time_format_option,
                route_short_name,
                schedule_type,
                direction_id
            )

            if df.empty:
                print(f"      No data to export for direction_id '{direction_id}'.")
                continue

            # Add DataFrame to the sheets dictionary with sheet name as 'Direction_{direction_id}'
            sheet_name = f"Direction_{direction_id}"
            df_sheets[sheet_name] = df

        if not df_sheets:
            print(f"    No data to export for route '{route_short_name}' with schedule '{schedule_type}'.")
            continue

        # Sanitize schedule_type for filename
        schedule_type_safe = schedule_type.replace(' ', '_').replace('-', '_').replace('/', '_')

        # Define output file path
        output_file = os.path.join(
            base_output_path,
            f"route_{route_short_name}_schedule_{schedule_type_safe}.xlsx"
        )

        # Export to Excel with multiple sheets
        export_to_excel_multiple_sheets(df_sheets, output_file)


