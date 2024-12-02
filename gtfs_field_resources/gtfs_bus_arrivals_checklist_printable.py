

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

# ==============================
# CONFIGURATION SECTION - CUSTOMIZE HERE
# ==============================

# Output directory
base_output_path = r'\\your_file_path\here\\'

# Define columns to read as strings
dtype_dict = {
    'stop_id': str,
    'trip_id': str,
    'route_id': str,
    'service_id': str,
    # Add other ID fields as needed
}

# Input file paths to load GTFS files with specified dtypes
base_input_path = r'\\your_file_path\here\\'

# List of required GTFS files
gtfs_files = ['trips.txt', 'stop_times.txt', 'routes.txt', 'stops.txt', 'calendar.txt']

# Check for existence of input directory
if not os.path.exists(base_input_path):
    raise FileNotFoundError(f"The input directory {base_input_path} does not exist.")

# Load GTFS files with specified dtypes
for file_name in gtfs_files:
    file_path = os.path.join(base_input_path, file_name)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The required GTFS file {file_name} does not exist in {base_input_path}.")

trips = pd.read_csv(os.path.join(base_input_path, 'trips.txt'), dtype=dtype_dict)
stop_times = pd.read_csv(os.path.join(base_input_path, 'stop_times.txt'), dtype=dtype_dict)
routes = pd.read_csv(os.path.join(base_input_path, 'routes.txt'), dtype=dtype_dict)
stops = pd.read_csv(os.path.join(base_input_path, 'stops.txt'), dtype=dtype_dict)
calendar = pd.read_csv(os.path.join(base_input_path, 'calendar.txt'), dtype=dtype_dict)

# Define clusters with stop IDs (e.g., bus centers with multiple nearby stops)
# Format: {'Cluster Name': ['stop_id1', 'stop_id2', ...]}
clusters = {
    'Your Cluster 1': ['1', '2', '3'],   # Replace with your cluster name and stop IDs
    'Your Cluster 2': ['4', '5', '6'],
    'Your Cluster 3': ['7', '8', '9', '10'],
    # Add more clusters as needed
}

# Define schedule types and corresponding days in the calendar
# Format: {'Schedule Type': ['day1', 'day2', ...]}
schedule_types = {
    'Weekday': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday'],
    'Saturday': ['saturday'],
    'Sunday': ['sunday'],
    # 'Friday': ['friday'],  # Uncomment if you have a unique Friday schedule
}

# Time windows for filtering trips
# Format: {'Schedule Type': {'Time Window Name': ('Start Time', 'End Time')}}
# Times should be in 'HH:MM:SS' 24-hour format
time_windows = {
    'Weekday': {
        'morning': ('06:00:00', '09:59:59'),
        'afternoon': ('14:00:00', '17:59:59'),
        # 'evening': ('18:00:00', '21:59:59'),  # Add as needed
    },
    'Saturday': {
        'midday': ('10:00:00', '13:59:59'),
        # Add more time windows for Saturday if needed
    },
    # 'Sunday': {  # Uncomment and customize for Sunday if needed
    #     'morning': ('08:00:00', '11:59:59'),
    #     'afternoon': ('12:00:00', '15:59:59'),
    # },
}

# ==============================
# END OF CONFIGURATION SECTION
# ==============================

# Create the output directory if it doesn't exist
if not os.path.exists(base_output_path):
    os.makedirs(base_output_path)

# Function to fix time format
def fix_time_format(time_str):
    parts = time_str.split(":")
    hours = int(parts[0])
    minutes = int(parts[1])
    seconds = int(parts[2])
    if hours >= 24:
        hours -= 24
    return f"{hours:02}:{minutes:02}:{seconds:02}"

# Ensure 'stop_id' is string in stops DataFrame
stops['stop_id'] = stops['stop_id'].astype(str)

# Process each schedule type
for schedule_name, days in schedule_types.items():
    print(f"Processing schedule: {schedule_name}")
    # Filter services for the current schedule type
    service_mask = calendar[days].astype(bool).all(axis=1)
    relevant_service_ids = calendar.loc[service_mask, 'service_id']

    # Filter trips to include only those that match the relevant service IDs
    trips_filtered = trips[trips['service_id'].isin(relevant_service_ids)]

    if trips_filtered.empty:
        print(f"No trips found for {schedule_name} schedule. Skipping.")
        continue

    # Merge trips with stop_times and routes to include route_short_name and block_id
    merged_data = pd.merge(stop_times, trips_filtered, on='trip_id')
    merged_data = pd.merge(merged_data, routes[['route_id', 'route_short_name']], on='route_id')

    # Ensure 'stop_id' is string in merged_data
    merged_data['stop_id'] = merged_data['stop_id'].astype(str)

    # Create a new column sequence_long
    merged_data['sequence_long'] = 'middle'

    # Assign "start" to sequence_long for rows with stop_sequence 1
    merged_data.loc[merged_data['stop_sequence'] == 1, 'sequence_long'] = 'start'

    # Get the highest stop_sequence number value for each trip
    max_sequence = merged_data.groupby('trip_id')['stop_sequence'].transform('max')

    # Assign "last" to sequence_long for rows with the highest stop_sequence number
    merged_data.loc[merged_data['stop_sequence'] == max_sequence, 'sequence_long'] = 'last'

    # Process each cluster
    for cluster_name, cluster_stop_ids in clusters.items():
        print(f"Processing cluster: {cluster_name} for {schedule_name} schedule")
        # Ensure cluster_stop_ids are strings
        cluster_stop_ids = [str(sid) for sid in cluster_stop_ids]

        # Filter merged_data by stop_id for the current cluster
        cluster_data = merged_data[merged_data['stop_id'].isin(cluster_stop_ids)]

        if cluster_data.empty:
            print(f"No data found for {cluster_name} on {schedule_name} schedule. Skipping.")
            continue

        # Apply the function to the time columns
        cluster_data['arrival_time'] = cluster_data['arrival_time'].apply(fix_time_format)
        cluster_data['departure_time'] = cluster_data['departure_time'].apply(fix_time_format)

        # Ensure times are strings to retain formatting in Excel
        cluster_data['arrival_time'] = cluster_data['arrival_time'].astype(str)
        cluster_data['departure_time'] = cluster_data['departure_time'].astype(str)

        # Convert arrival_time and departure_time to datetime
        cluster_data['arrival_time'] = pd.to_datetime(cluster_data['arrival_time'], format='%H:%M:%S').dt.time
        cluster_data['departure_time'] = pd.to_datetime(cluster_data['departure_time'], format='%H:%M:%S').dt.time

        # Sort by arrival_time
        cluster_data = cluster_data.sort_values(by='arrival_time')

        # Add 'act_arrival' and 'act_departure' columns with placeholders
        cluster_data.insert(cluster_data.columns.get_loc('arrival_time') + 1, 'act_arrival', '________')
        cluster_data.insert(cluster_data.columns.get_loc('departure_time') + 1, 'act_departure', '________')

        # Modify 'act_arrival' where 'sequence_long' is 'start'
        cluster_data.loc[cluster_data['sequence_long'] == 'start', 'act_arrival'] = '__XXXX__'

        # Modify 'act_departure' where 'sequence_long' is 'last'
        cluster_data.loc[cluster_data['sequence_long'] == 'last', 'act_departure'] = '__XXXX__'

        # Add 'bus_number' column with underscores
        cluster_data['bus_number'] = '________'

        # Add 'stop_name' column next to 'stop_id'
        cluster_data = pd.merge(cluster_data, stops[['stop_id', 'stop_name']], on='stop_id', how='left')

        # Move specified columns to desired positions
        first_columns = [
            'route_short_name', 'trip_headsign', 'stop_sequence', 'sequence_long',
            'stop_id', 'stop_name', 'arrival_time', 'act_arrival',
            'departure_time', 'act_departure', 'block_id', 'bus_number'
        ]
        other_columns = [col for col in cluster_data.columns if col not in first_columns]
        cluster_data = cluster_data[first_columns + other_columns]

        # Drop unnecessary columns
        cluster_data = cluster_data.drop(columns=[
            'shape_dist_traveled', 'shape_id', 'route_id', 'service_id',
            'trip_id', 'timepoint', 'direction_id', 'stop_headsign', 'pickup_type',
            'drop_off_type', 'wheelchair_accessible', 'bikes_allowed', 'trip_short_name'
        ], errors='ignore')

        # Define the output file name for all trips
        output_file_name = f'{cluster_name}_{schedule_name}_data.xlsx'
        output_file = os.path.join(base_output_path, output_file_name)

        # Export all cluster data to Excel with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            cluster_data.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Align all headers to the left
            for cell in worksheet[1]:
                cell.alignment = Alignment(horizontal='left')

            # Adjust the width of all columns
            for idx, col in enumerate(cluster_data.columns, 1):  # 1-based indexing for Excel columns
                column_letter = get_column_letter(idx)
                max_length = max(
                    cluster_data[col].astype(str).map(len).max(),  # Maximum length of column entries
                    len(str(col))  # Length of the column header
                ) + 2  # Adding extra space for better readability
                worksheet.column_dimensions[column_letter].width = max_length

        print(f"Processed and exported data for {cluster_name} on {schedule_name} schedule.")

        # Now, check if there are time windows for this schedule
        if schedule_name in time_windows:
            for time_window_name, time_range in time_windows[schedule_name].items():
                start_time_str, end_time_str = time_range
                start_time = pd.to_datetime(start_time_str, format='%H:%M:%S').time()
                end_time = pd.to_datetime(end_time_str, format='%H:%M:%S').time()

                # Filter cluster_data to only include trips within the time window
                filtered_data = cluster_data[
                    (cluster_data['arrival_time'] >= start_time) &
                    (cluster_data['arrival_time'] <= end_time)
                ]

                if filtered_data.empty:
                    print(f"No data found for {cluster_name} on {schedule_name} schedule in {time_window_name} time window. Skipping.")
                    continue

                # Define the output file name for the time window
                output_file_name = f'{cluster_name}_{schedule_name}_{time_window_name}_data.xlsx'
                output_file = os.path.join(base_output_path, output_file_name)

                # Export filtered data to Excel with formatting
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    filtered_data.to_excel(writer, index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']

                    # Align all headers to the left
                    for cell in worksheet[1]:
                        cell.alignment = Alignment(horizontal='left')

                    # Adjust the width of all columns
                    for idx, col in enumerate(filtered_data.columns, 1):  # 1-based indexing for Excel columns
                        column_letter = get_column_letter(idx)
                        max_length = max(
                            filtered_data[col].astype(str).map(len).max(),  # Maximum length of column entries
                            len(str(col))  # Length of the column header
                        ) + 2  # Adding extra space for better readability
                        worksheet.column_dimensions[column_letter].width = max_length

                print(f"Processed and exported data for {cluster_name} on {schedule_name} schedule in {time_window_name} time window.")

print("All clusters and schedules have been processed and exported.")



