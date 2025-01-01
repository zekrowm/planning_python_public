"""
gtfs_stop_boardings_by_route_printable

This module processes GTFS data to generate printable boarding schedules
by route, applying specified time windows and formatting the output in Excel.
"""

import os
import sys

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ==============================
# CONFIGURATION SECTION
# ==============================

# Input file paths for GTFS files
BASE_INPUT_PATH = r'C:\Your\Path\To\GTFS_folder'
TRIPS_FILE = os.path.join(BASE_INPUT_PATH, "trips.txt")
STOP_TIMES_FILE = os.path.join(BASE_INPUT_PATH, "stop_times.txt")
ROUTES_FILE = os.path.join(BASE_INPUT_PATH, "routes.txt")
STOPS_FILE = os.path.join(BASE_INPUT_PATH, "stops.txt")
CALENDAR_FILE = os.path.join(BASE_INPUT_PATH, "calendar.txt")

# Output directory
BASE_OUTPUT_PATH = r'C:\Your\Path\To\Output_folder'
if not os.path.exists(BASE_OUTPUT_PATH):
    os.makedirs(BASE_OUTPUT_PATH)

# Define schedule types and corresponding days in the calendar
# Format: {'Schedule Type': ['day1', 'day2', ...]}
SCHEDULE_TYPES = {
    'Weekday': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday'],
    'Saturday': ['saturday'],
    'Sunday': ['sunday'],
    # 'Friday': ['friday'],  # Uncomment if you have a unique Friday schedule
}

# Time windows for filtering trips
# Format: {'Schedule Type': {'Time Window Name': ('Start Time', 'End Time')}}
# Times should be in 'HH:MM' 24-hour format
TIME_WINDOWS = {
    'Weekday': {
        'full_day': ('00:00', '23:59'),
        'morning': ('06:00', '09:59'),
        'afternoon': ('14:00', '17:59'),
        # 'evening': ('18:00', '21:59'),  # Add as needed
    },
    'Saturday': {
        'midday': ('10:00', '13:59'),
        # Add more time windows for Saturday if needed
    },
    # 'Sunday': {  # Uncomment and customize for Sunday if needed
    #     'morning': ('08:00', '11:59'),
    #     'afternoon': ('12:00', '15:59'),
    # },
}

# Placeholder values
MISSING_TIME = "________"
COMMENTS_PLACEHOLDER = "__________________"
MAX_COLUMN_WIDTH = 50

# ==============================
# END OF CONFIGURATION
# ==============================


def load_data():
    """
    Loads GTFS data from specified input files.

    Returns:
        tuple: DataFrames containing trips, stop_times, routes, stops, and calendar data.
    """
    try:
        trips = pd.read_csv(TRIPS_FILE)
        stop_times = pd.read_csv(STOP_TIMES_FILE)
        routes = pd.read_csv(ROUTES_FILE)
        stops = pd.read_csv(STOPS_FILE)
        calendar = pd.read_csv(CALENDAR_FILE)

        # If 'timepoint' column does not exist, assume all stops are timepoints
        if 'timepoint' not in stop_times.columns:
            stop_times['timepoint'] = 1
    except FileNotFoundError as e:
        print(f"File not found: {e}")
        sys.exit(1)
    except pd.errors.ParserError as e:
        print(f"Error parsing CSV files: {e}")
        sys.exit(1)
    return trips, stop_times, routes, stops, calendar


def time_to_minutes(time_str):
    """
    Converts a time string in 'HH:MM:SS' or 'HH:MM' format to total minutes since midnight.

    Args:
        time_str (str): Time string to convert.

    Returns:
        float or None: Total minutes since midnight, or None if format is incorrect.
    """
    parts = time_str.split(":")
    if len(parts) == 3:
        try:
            hours, minutes, seconds = map(int, parts)
        except ValueError:
            return None
    elif len(parts) == 2:
        try:
            hours, minutes = map(int, parts)
            seconds = 0
        except ValueError:
            return None
    else:
        return None
    return hours * 60 + minutes + seconds / 60


def hhmm_to_minutes(time_str):
    """
    Converts a time string in 'HH:MM' format to total minutes since midnight.

    Args:
        time_str (str): Time string to convert.

    Returns:
        int or None: Total minutes since midnight, or None if format is incorrect.
    """
    parts = time_str.split(":")
    if len(parts) == 2:
        try:
            hours, minutes = map(int, parts)
            return hours * 60 + minutes
        except ValueError:
            return None
    else:
        return None


def format_time(time_str):
    """
    Formats a time string to 'HH:MM' format, handling times over 24 hours.

    Args:
        time_str (str): Time string to format.

    Returns:
        str: Formatted time string.
    """
    total_minutes = time_to_minutes(time_str)
    if total_minutes is not None:
        hours = int(total_minutes // 60)
        minutes = int(total_minutes % 60)
        return f"{hours:02}:{minutes:02}"
    else:
        return time_str


def export_to_excel(df, output_file, bold_rows=None):
    """
    Exports a DataFrame to an Excel file with formatting.

    Args:
        df (pd.DataFrame): DataFrame to export.
        output_file (str): Path to the output Excel file.
        bold_rows (list, optional): List of row indices to bold.
    """
    if df.empty:
        print(f"No data to export to {output_file}.")
        return
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Schedule')
        sheet = writer.sheets['Schedule']

        # Format headers and columns
        for col_num, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_num)
            try:
                max_length = max(
                    (len(str(value)) for value in df[col_name]),
                    default=0
                )
                max_length = max(max_length, len(col_name))
                sheet.column_dimensions[col_letter].width = min(
                    max_length + 2, MAX_COLUMN_WIDTH
                )
            except Exception as e:
                print(f"Error setting column width for {col_name}: {e}")

        # Bold rows where Actual Time == MISSING_TIME
        if bold_rows is not None:
            bold_font = Font(bold=True)
            for row_idx in bold_rows:
                for col_num in range(1, len(df.columns) + 1):
                    cell = sheet.cell(row=row_idx + 2, column=col_num)  # +2 for header and 1-indexing
                    cell.font = bold_font


# Load GTFS data
TRIPS, STOP_TIMES, ROUTES, STOPS, CALENDAR = load_data()

# Extract unique route short names
ROUTE_SHORT_NAMES = ROUTES['route_short_name'].unique()


# Generate schedule for each configuration
def generate_schedules():
    """
    Generates and exports boarding schedules based on GTFS data and configurations.
    """
    for schedule_type, days in SCHEDULE_TYPES.items():
        relevant_services = CALENDAR[
            CALENDAR[days].all(axis=1)
        ]['service_id']
        for route_short_name in ROUTE_SHORT_NAMES:
            route_ids = ROUTES[
                ROUTES['route_short_name'] == route_short_name
            ]['route_id']
            relevant_trips = TRIPS[
                (TRIPS['route_id'].isin(route_ids)) &
                (TRIPS['service_id'].isin(relevant_services))
            ]
            for direction_id in relevant_trips['direction_id'].unique():
                trips_in_direction = relevant_trips[
                    relevant_trips['direction_id'] == direction_id
                ]
                timepoints = STOP_TIMES[
                    STOP_TIMES['trip_id'].isin(trips_in_direction['trip_id'])
                ].copy()
                timepoints['departure_minutes'] = timepoints[
                    'departure_time'
                ].apply(time_to_minutes)
                timepoints['formatted_departure_time'] = timepoints[
                    'departure_time'
                ].apply(format_time)

                if schedule_type in TIME_WINDOWS:
                    for time_window_name, (start_time, end_time) in TIME_WINDOWS[
                        schedule_type
                    ].items():
                        start_minutes = hhmm_to_minutes(start_time)
                        end_minutes = hhmm_to_minutes(end_time)

                        timepoints_filtered = timepoints[
                            (timepoints['departure_minutes'] >= start_minutes) &
                            (timepoints['departure_minutes'] <= end_minutes)
                        ]

                        if not timepoints_filtered.empty:
                            # Merge with trips to get route_id, service_id, direction_id
                            timepoints_filtered = timepoints_filtered.merge(
                                TRIPS[
                                    [
                                        'trip_id', 'route_id',
                                        'service_id', 'trip_headsign',
                                        'direction_id'
                                    ]
                                ],
                                on='trip_id',
                                how='left'
                            )

                            # Merge with stops to get stop_name
                            timepoints_filtered = timepoints_filtered.merge(
                                STOPS[['stop_id', 'stop_name']],
                                on='stop_id',
                                how='left'
                            )

                            # Merge with routes to get route_short_name
                            timepoints_filtered = timepoints_filtered.merge(
                                ROUTES[['route_id', 'route_short_name']],
                                on='route_id',
                                how='left'
                            )

                            # Create 'Trip ID + Start Time'
                            # For each trip_id, get the first 'departure_time'
                            first_departure_times = (
                                timepoints_filtered
                                .groupby('trip_id')['departure_minutes']
                                .min()
                                .reset_index()
                            )
                            first_departure_times.rename(
                                columns={'departure_minutes': 'trip_start_minutes'},
                                inplace=True
                            )
                            timepoints_filtered = timepoints_filtered.merge(
                                first_departure_times,
                                on='trip_id',
                                how='left'
                            )

                            timepoints_filtered['trip_start_time'] = (
                                timepoints_filtered['trip_start_minutes']
                                .apply(lambda x: f"{int(x // 60):02}:{int(x % 60):02}")
                            )

                            timepoints_filtered['Trip ID + Start Time'] = (
                                timepoints_filtered['trip_id'].astype(str) + ' ' +
                                timepoints_filtered['trip_start_time']
                            )

                            # Map 'direction_id' to 'Direction' (e.g., 0 to 'Outbound', 1 to 'Inbound')
                            direction_mapping = {0: 'Outbound', 1: 'Inbound'}
                            timepoints_filtered['Direction'] = timepoints_filtered[
                                'direction_id'
                            ].map(direction_mapping)

                            # Add 'Service Period' as 'schedule_type'
                            timepoints_filtered['Service Period'] = schedule_type

                            # Add 'Route' from 'route_short_name'
                            timepoints_filtered['Route'] = timepoints_filtered[
                                'route_short_name'
                            ]

                            # Create 'Scheduled Time' from 'formatted_departure_time'
                            timepoints_filtered['Scheduled Time'] = timepoints_filtered[
                                'formatted_departure_time'
                            ]

                            # Create 'Actual Time' column
                            timepoints_filtered['Actual Time'] = timepoints_filtered[
                                'timepoint'
                            ].apply(
                                lambda x: MISSING_TIME if x == 1 else 'X'
                            )

                            # 'Stop Sequence' from 'stop_sequence'
                            timepoints_filtered['Stop Sequence'] = (
                                timepoints_filtered['stop_sequence']
                            )

                            # 'Stop ID' from 'stop_id'
                            timepoints_filtered['Stop ID'] = (
                                timepoints_filtered['stop_id']
                            )

                            # 'Stop Name' from 'stop_name'
                            timepoints_filtered['Stop Name'] = (
                                timepoints_filtered['stop_name']
                            )

                            # 'Date' as placeholder
                            timepoints_filtered['Date'] = MISSING_TIME

                            # Determine first and last stops for 'On' and 'Off' columns
                            first_stops = (
                                timepoints_filtered
                                .groupby('trip_id')['stop_sequence']
                                .min()
                                .reset_index()
                            )
                            first_stops.rename(
                                columns={'stop_sequence': 'first_stop_sequence'},
                                inplace=True
                            )
                            last_stops = (
                                timepoints_filtered
                                .groupby('trip_id')['stop_sequence']
                                .max()
                                .reset_index()
                            )
                            last_stops.rename(
                                columns={'stop_sequence': 'last_stop_sequence'},
                                inplace=True
                            )
                            timepoints_filtered = timepoints_filtered.merge(
                                first_stops, on='trip_id', how='left'
                            )
                            timepoints_filtered = timepoints_filtered.merge(
                                last_stops, on='trip_id', how='left'
                            )

                            # 'On' column
                            timepoints_filtered['On'] = timepoints_filtered.apply(
                                lambda row: 'X' if row['stop_sequence'] == row[
                                    'last_stop_sequence'
                                ] else MISSING_TIME,
                                axis=1
                            )

                            # 'Off' column
                            timepoints_filtered['Off'] = timepoints_filtered.apply(
                                lambda row: 'X' if row['stop_sequence'] == row[
                                    'first_stop_sequence'
                                ] else MISSING_TIME,
                                axis=1
                            )

                            # Select the required columns in the required order
                            columns_order = [
                                'Date', 'Service Period', 'Route', 'Direction',
                                'Trip ID + Start Time', 'Scheduled Time',
                                'Actual Time', 'Stop Sequence', 'Stop ID',
                                'Stop Name', 'On', 'Off'
                            ]

                            # Prepare the output DataFrame
                            output_df = timepoints_filtered[columns_order].copy()

                            # Sort the DataFrame
                            output_df['trip_start_minutes'] = (
                                timepoints_filtered['trip_start_minutes']
                            )
                            output_df = output_df.sort_values(
                                by=['trip_start_minutes', 'Trip ID + Start Time', 'Stop Sequence']
                            ).reset_index(drop=True)

                            # Remove the temporary 'trip_start_minutes' column
                            output_df = output_df.drop(columns=['trip_start_minutes'])

                            # Get indices of rows where 'Actual Time' == MISSING_TIME for bold formatting
                            bold_rows = output_df[
                                output_df['Actual Time'] == MISSING_TIME
                            ].index.tolist()

                            # Create a filename for the output Excel file
                            filename = (
                                f"{route_short_name}_{schedule_type}_"
                                f"{time_window_name}_dir_{direction_id}.xlsx"
                            )
                            output_file = os.path.join(BASE_OUTPUT_PATH, filename)

                            # Export to Excel
                            export_to_excel(
                                output_df,
                                output_file,
                                bold_rows=bold_rows
                            )
                else:
                    print(
                        f"No time windows specified for schedule type "
                        f"'{schedule_type}'. Skipping."
                    )


if __name__ == "__main__":
    generate_schedules()
