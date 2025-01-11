"""
Ridership by Route and Stop Processor

This script processes ridership data from an input Excel file by filtering specific routes and stop IDs,
aggregating the data for defined time periods, and exporting the results to a new Excel file with multiple
formatted sheets.
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# ==========================
# Configuration Section
# ==========================

# Input and Output File Paths
INPUT_FILE_PATH = r'\\Your\File\Path\to\ridership_by_route_and_stop_all_time_periods.xlsx'
OUTPUT_FILE_SUFFIX = '_processed'
OUTPUT_FILE_EXTENSION = '.xlsx'

# Routes and Stop IDs for Filtering
ROUTES = ["101", "202", "303"] # Replace with your route names
STOP_IDS = [
    1001, 1002, 1003, 1004
] # Replace with your stops of interest

# Required Columns in the Input Excel File
REQUIRED_COLUMNS = ['TIME_PERIOD', 'ROUTE_NAME', 'STOP', 'STOP_ID', 'BOARD_ALL', 'ALIGHT_ALL']

# Columns to Retain in the Output Sheets
COLUMNS_TO_RETAIN = ['ROUTE_NAME', 'STOP', 'STOP_ID', 'BOARD_ALL', 'ALIGHT_ALL']

# Time Periods for Aggregation
TIME_PERIODS = ['AM PEAK', 'PM PEAK'] # Replace with your time periods of interest

# ==========================
# End of Configuration
# ==========================

def main():
    # Define the input file path from the configuration
    input_file = INPUT_FILE_PATH

    # Define the output file path by adding '_processed' before the file extension
    base, ext = os.path.splitext(input_file)
    ext = ext.lower()

    # Ensure the output has the correct extension
    if ext != OUTPUT_FILE_EXTENSION:
        print(f"Warning: The input file has an unexpected extension '{ext}'. The output file will use '{OUTPUT_FILE_EXTENSION}' extension.")
        ext = OUTPUT_FILE_EXTENSION

    output_file = f"{base}{OUTPUT_FILE_SUFFIX}{ext}"

    # Check if the input file exists
    if not os.path.exists(input_file):
        print(f"Error: The file '{input_file}' does not exist.")
        exit(1)

    # Read the Excel file into a pandas DataFrame
    try:
        df = pd.read_excel(input_file)
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        exit(1)

    # Verify that the required columns are present
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]

    if missing_columns:
        print(f"Error: The following required columns are missing in the data: {missing_columns}")
        exit(1)

    # Filter the DataFrame based on ROUTES and STOP_IDS
    df_filtered = df[
        df['ROUTE_NAME'].isin(ROUTES) &
        df['STOP_ID'].isin(STOP_IDS)
    ]

    # Define the columns to retain in the new sheets
    columns_to_retain = COLUMNS_TO_RETAIN

    # Standardize the 'TIME_PERIOD' values by stripping whitespace and converting to uppercase
    df_filtered['TIME_PERIOD'] = df_filtered['TIME_PERIOD'].astype(str).str.strip().str.upper()

    # Filter data for each specified time period
    peak_data = {}
    for period in TIME_PERIODS:
        peak_data[period] = df_filtered[df_filtered['TIME_PERIOD'] == period][columns_to_retain]

    # Function to aggregate data by STOP and STOP_ID and create ROUTES column
    def aggregate_by_stop(df_subset):
        # Group by STOP and STOP_ID
        aggregated = df_subset.groupby(['STOP', 'STOP_ID'], as_index=False).agg({
            'BOARD_ALL': 'sum',
            'ALIGHT_ALL': 'sum',
            # Collect unique ROUTE_NAMEs into a comma-separated string
            'ROUTE_NAME': lambda x: ', '.join(sorted(x.unique()))
        })
        # Rename columns to indicate they are totals
        aggregated.rename(columns={
            'BOARD_ALL': 'BOARD_ALL_TOTAL',
            'ALIGHT_ALL': 'ALIGHT_ALL_TOTAL',
            'ROUTE_NAME': 'ROUTES'
        }, inplace=True)
        return aggregated

    # Aggregate data for All Time Periods
    all_time_aggregated = aggregate_by_stop(df_filtered)

    # Aggregate data for each specified time period
    aggregated_peaks = {}
    for period, data in peak_data.items():
        aggregated_peaks[period] = aggregate_by_stop(data)

    # Round ridership columns to 1 decimal place
    ridership_columns = ['BOARD_ALL_TOTAL', 'ALIGHT_ALL_TOTAL']
    for df_agg in [all_time_aggregated] + list(aggregated_peaks.values()):
        for col in ridership_columns:
            if col in df_agg.columns:
                df_agg[col] = df_agg[col].round(1)

    # Also round the original dataframe's ridership columns if needed
    for col in ['BOARD_ALL', 'ALIGHT_ALL']:
        if col in df_filtered.columns:
            df_filtered[col] = df_filtered[col].round(1)

    # Write the data to a new Excel file with multiple sheets and adjust column widths
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write the original filtered data to a sheet named 'Original'
            df_filtered.to_excel(writer, sheet_name='Original', index=False)

            # Write the aggregated data for each time period
            for period, df_agg in aggregated_peaks.items():
                df_agg.to_excel(writer, sheet_name=period, index=False)

            # Write the All Time Periods aggregated data to a new sheet
            all_time_aggregated.to_excel(writer, sheet_name='All Time Periods', index=False)

            # Save the writer to ensure data is written before adjusting column widths
            writer.save()

        # Now, open the workbook using openpyxl to adjust column widths and format headers
        workbook = load_workbook(output_file)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Bold the header row
            for cell in sheet[1]:
                cell.font = Font(bold=True)

            for column_cells in sheet.columns:
                # Get the maximum length of the content in the column
                max_length = 0
                column = column_cells[0].column_letter  # Get the column name (e.g., 'A', 'B', ...)
                for cell in column_cells:
                    try:
                        cell_value = str(cell.value)
                        if cell_value is None:
                            cell_length = 0
                        else:
                            cell_length = len(cell_value)
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                # Set the column width with a little extra space
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook after adjusting column widths and formatting
        workbook.save(output_file)

        print(f"Success: The processed file has been saved as '{output_file}'.")
    except Exception as e:
        print(f"Error writing the processed Excel file: {e}")
        exit(1)

if __name__ == "__main__":
    main()
